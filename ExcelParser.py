import os

import openpyxl

# os.chdir('') #write path where excel doc is, if it isn't in cwd

# --------------------
# read excel files
# --------------------

workbook = openpyxl.load_workbook("example.xlsx")
print(type(workbook))

sheet = workbook["Sheet1"]
print(type(sheet))

print(workbook.sheetnames)

print((sheet["A1"].value))

# another method to read the data within cells
print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

