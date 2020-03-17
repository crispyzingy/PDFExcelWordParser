import openpyxl

import os

# -------------------------
# write/edit excel files
# -------------------------
wb = openpyxl.Workbook()
print(wb.sheetnames)

sheet = wb["Sheet"]  # use the sheet called 'Sheet'

print(sheet["A1"].value == None)

sheet["A1"] = 42
sheet["A2"] = "Hello"

# os.chdir("c:\\Users\\crisp\\Documents") # choose new directory or it'll save at cwd
wb.save("eg2.xlsx")
sheet2 = wb.create_sheet()
print(wb.sheetnames)
sheet2.title = "My new Sheet Name"
print(wb.sheetnames)
wb.save("eg3.xlsx")
wb.create_sheet(index=0, title="My Other Sheet")  # specify poistion of sheet
wb.save("eg3.xlsx")

